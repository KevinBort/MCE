import openpyxl, os
import re
from openpyxl.styles import Alignment
# Modulo 1 -----> Generar la ruta hacia el archivo a modificar.
script_dir = os.getcwd()
template_dir = os.path.join(script_dir, 'CEM_Template.xlsx')

# Modulo 2 -----> Generar un objeto de Excel con Openpyxl
if os.path.exists(template_dir):
    wb = openpyxl.load_workbook(template_dir)
    ws = wb.active
else:
    print('File not found')

causas = [] # ACA SE VAN A GUARDAR LAS TUPLAS EN FORMA DE (PREFIX, DESCRIPTION)
acciones = []

suffix_dic = {'AH': 'High', 'SHH': 'High High', 'AL': 'Low', 'SLL': 'Low Low'}

# Modulo 3 -----> Input de Excel


input_file_name = 'datos_entrada.xlsx'
input_file_path = os.path.join(script_dir, input_file_name)

wb_input = openpyxl.load_workbook(input_file_name)
ws_input = wb_input.active
# ---------- I N P U T --------- T I T U L O   Y   C O D I G O   ----------
title = ws_input.cell(row=1, column=3).value
name = ws_input.cell(row=2, column=3).value
code = ws_input.cell(row=3, column=3).value 
# ---------- I N P U T --------- C A U S A S ----------
for row in ws_input.iter_rows(min_row=10, max_row=50, min_col=2, max_col=3, values_only=True):
    if row != (None, None):
        causas.append(row)
# ---------- I N P U T --------- A C C I O N E S  ----------
for row in ws_input.iter_rows(min_row=10, max_row=50, min_col=4, max_col=5, values_only=True):
    if row != (None, None):
        acciones.append(row)

wb.save(input_file_name) 
wb.close()       
col_index_acciones=11

for actuador,description_act in acciones:
    ws.cell(row=2, column=col_index_acciones + 1, value=description_act)
    ws.cell(row=3, column=col_index_acciones + 1 ,value=actuador)
    col_index_acciones+=1


# Modulo 4 -----> Definicion de los prefijos que son parte del string que está en la posición 0 de la tupla ingresada. 
prefix_descriptions = {'TT': 'Temperature',
                       'PT': 'Pressure',
                       'LT': 'Level',
                       'FT': 'Flow'}

# Inicialización del índice de filas, es a partir de donde va a empezar a escribir el código. 
row_index_causas = 5

for prefix, description in causas: 
    # Extracción del número de 4 dígitos con regex
    match = re.search(r'-(\d{4})', prefix)
    if match:
        number = match.group(1)
    else:
        number = ''

    # Partiendo del prefijo, TT, PT, etc., se arma lo que sería la descripción del instrumento, y se setea el valor unknown por defecto en caso de
    # que la llave no exista, o se escriba mal.

    prefix_description = prefix_descriptions.get(prefix[:2], 'Unknown')

    for key, value in suffix_dic.items():    
        # Create the updated description
        updated_description = f' {value} {description}' # suffix_dic={'AH':'High', 'SHH':'High High', 'AL':'Low', 'SLL':'Low Low'}

        formatted_string = f'{prefix[:1]}{key}-{number}'  # Include the extracted 4-digit number
       
        ws.cell(row=row_index_causas + 1, column=4, value=prefix) # Escribe la fila del TAG.
        ws.cell(row_index_causas+ 1, column=5, value=updated_description)  # Escribe la descripcion del instrumento
        ws.cell(row_index_causas + 1, column=7, value=formatted_string)  # Escribe el TAG de aviso o trip
        row_index_causas += 1
ws.cell(row=40,column=6,value=code)

ws.cell(row=42,column=6,value=name)
ws.cell(row=2,column=3,value=title)
merge_range='C2:I3'
ws.merge_cells(merge_range)



# Save the updated workbook (overwrite the template file)
wb.save('CEM_mod_v1.xlsx')

#   R E      H A C E R     E L    A  R  C  H  I  V  O  E  N T R A D A 